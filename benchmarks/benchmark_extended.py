#!/usr/bin/env python3
"""Lookup/formula benchmark scenarios using the common 10K x 15 workload."""

import sys
import tempfile
import time
from pathlib import Path

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "bindings" / "python"))
import xlpp

import openpyxl
import xlsxwriter

ROWS = 10_000
COLS = 15
LOOKUP_ROWS = 500


def value(row, col):
    if col == 0:
        return f"Item-{row}"
    if col == 1:
        return row + 1
    if col == 2:
        return round((row * 7919 % 10_000_000) / 100.0, 2)
    return "Lorem ipsum dolor sit amet"


def cells(sheet, row, formula=False):
    for col in range(COLS):
        if formula and col == COLS - 1:
            sheet.cell(row, col + 1).set_formula(
                f'=VLOOKUP(A{row},Lookup!$A$1:$B${LOOKUP_ROWS},2,FALSE)'
            )
        else:
            sheet.cell(row, col + 1).value = value(row - 1, col)


def write_xlpp(path, scenario):
    workbook = xlpp.Workbook()
    data = workbook.add_worksheet("Data")
    lookup = workbook.add_worksheet("Lookup")
    for row in range(1, LOOKUP_ROWS + 1):
        lookup.cell(row, 1).value = f"Item-{row - 1}"
        lookup.cell(row, 2).value = row * 1.25
    for row in range(1, ROWS + 1):
        cells(data, row, scenario == "formula")
    workbook.save(str(path))


def write_openpyxl(path, scenario):
    workbook = openpyxl.Workbook()
    data = workbook.active
    data.title = "Data"
    lookup = workbook.create_sheet("Lookup")
    for row in range(1, LOOKUP_ROWS + 1):
        lookup.cell(row, 1, f"Item-{row - 1}")
        lookup.cell(row, 2, row * 1.25)
    for row in range(1, ROWS + 1):
        end_col = COLS - 1 if scenario == "formula" else COLS
        data.append([f"Item-{row - 1}", row, value(row - 1, 2), *[value(row - 1, c) for c in range(3, end_col)]])
        if scenario == "formula":
            data.cell(row, COLS).value = f'=VLOOKUP(A{row},Lookup!$A$1:$B${LOOKUP_ROWS},2,FALSE)'
    workbook.save(path)


def write_xlsxwriter(path, scenario):
    workbook = xlsxwriter.Workbook(str(path))
    data = workbook.add_worksheet("Data")
    lookup = workbook.add_worksheet("Lookup")
    for row in range(LOOKUP_ROWS):
        lookup.write(row, 0, f"Item-{row}")
        lookup.write_number(row, 1, (row + 1) * 1.25)
    for row in range(ROWS):
        for col in range(COLS):
            if scenario == "formula" and col == COLS - 1:
                data.write_formula(row, col, f'=VLOOKUP(A{row + 1},Lookup!$A$1:$B${LOOKUP_ROWS},2,FALSE)')
            else:
                data.write(row, col, value(row, col))
    workbook.close()


def read_xlpp(path):
    workbook = xlpp.Workbook()
    workbook.load(str(path))
    return workbook["Data"].max_row


def read_openpyxl(path):
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    count = sum(1 for _ in workbook["Data"].iter_rows())
    workbook.close()
    return count


def run(scenario, library, writer, reader):
    path = Path(tempfile.gettempdir()) / f"xlpp-{library}-{scenario}.xlsx"
    start = time.perf_counter()
    writer(path, scenario)
    write_ms = (time.perf_counter() - start) * 1000
    start = time.perf_counter()
    reader(path)
    read_ms = (time.perf_counter() - start) * 1000
    print(f"BENCHMARK,python,{library},{scenario}_write,{write_ms:.2f},{path.stat().st_size}")
    print(f"BENCHMARK,python,{library},{scenario}_read,{read_ms:.2f},0")
    path.unlink(missing_ok=True)


def run_write_only(scenario, library, writer):
    path = Path(tempfile.gettempdir()) / f"xlpp-{library}-{scenario}.xlsx"
    start = time.perf_counter()
    writer(path, scenario)
    write_ms = (time.perf_counter() - start) * 1000
    print(f"BENCHMARK,python,{library},{scenario}_write,{write_ms:.2f},{path.stat().st_size}")
    path.unlink(missing_ok=True)


for scenario in ("lookup", "formula"):
    run(scenario, "XLPP", write_xlpp, read_xlpp)
    run(scenario, "openpyxl", write_openpyxl, read_openpyxl)
    run_write_only(scenario, "XlsxWriter", write_xlsxwriter)
