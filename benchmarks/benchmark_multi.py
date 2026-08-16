import time
import statistics
import tempfile
import os

import xlpp
import openpyxl
import xlsxwriter

ROUNDS = 5
CONFIGS = [
    (500, 200, "100K cells (500x200)"),
    (700, 300, "210K cells (700x300)"),
]

OUT = os.path.join(tempfile.gettempdir(), "bench_out")


def build_row(r, cols):
    row = []
    for c in range(cols):
        if c % 4 == 0:
            row.append(r * 1000 + c)
        elif c % 4 == 1:
            row.append((r * 1000 + c) / 7.0)
        elif c % 4 == 2:
            row.append(f"text-{r}-{c}")
        else:
            row.append(True)
    return row


def run_xlpp_write(rows, cols):
    wb = xlpp.Workbook()
    ws = wb.add_worksheet("Data")
    for r in range(rows):
        ws.append(build_row(r, cols))
    wb.save(OUT + "_xlpp.xlsx")
    return wb


def run_xlpp_read(rows, cols):
    wb = xlpp.Workbook()
    wb.load(OUT + "_xlpp.xlsx")
    ws = wb.worksheet("Data")
    total = 0
    for r in range(1, rows + 1):
        for c in range(1, cols + 1):
            v = ws.try_cell(r, c)
            if v is not None:
                total += 1
    return total


def run_openpyxl_write(rows, cols):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    for r in range(rows):
        ws.append(build_row(r, cols))
    wb.save(OUT + "_op.xlsx")


def run_openpyxl_read(rows, cols):
    wb = openpyxl.load_workbook(OUT + "_op.xlsx", read_only=False)
    ws = wb["Data"]
    count = 0
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if v is not None:
                count += 1


def run_xlsxwriter_write(rows, cols):
    wb = xlsxwriter.Workbook(OUT + "_xw.xlsx", {"constant_memory": False})
    ws = wb.add_worksheet("Data")
    for r in range(rows):
        for c in range(cols):
            if c % 4 == 0:
                ws.write_number(r, c, r * 1000 + c)
            elif c % 4 == 1:
                ws.write_number(r, c, (r * 1000 + c) / 7.0)
            elif c % 4 == 2:
                ws.write_string(r, c, f"text-{r}-{c}")
            else:
                ws.write_boolean(r, c, True)
    wb.close()


def bench(label, fn, args):
    samples = []
    for _ in range(ROUNDS):
        t0 = time.perf_counter()
        fn(*args)
        samples.append((time.perf_counter() - t0) * 1000.0)
    avg = statistics.mean(samples)
    sd = statistics.stdev(samples) if len(samples) > 1 else 0.0
    print(f"  {label:<42} avg={avg:8.1f} ms  (runs: {', '.join(f'{s:.1f}' for s in samples)})")
    return avg


print(f"Rounds per test: {ROUNDS}\n")
for rows, cols, label in CONFIGS:
    cells = rows * cols
    print(f"=== {label} ===")
    print(f"[write] XL++ (Python)")
    bench("xlpp", run_xlpp_write, (rows, cols))
    print(f"[write] openpyxl")
    bench("openpyxl", run_openpyxl_write, (rows, cols))
    print(f"[write] xlsxwriter")
    bench("xlsxwriter", run_xlsxwriter_write, (rows, cols))
    print(f"[read] XL++ (Python)")
    bench("xlpp", run_xlpp_read, (rows, cols))
    print(f"[read] openpyxl")
    bench("openpyxl", run_openpyxl_read, (rows, cols))
    print()
