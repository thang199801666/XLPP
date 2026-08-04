#!/usr/bin/env python3
"""Read large xlsx files with openpyxl's read-only row iterator."""

import tempfile
import time
from pathlib import Path

import openpyxl
import xlsxwriter


def generate(path, rows):
    workbook = xlsxwriter.Workbook(str(path), {"constant_memory": True})
    sheet = workbook.add_worksheet("Data")
    for row in range(rows):
        sheet.write_row(row, 0, [f"Item-{row}", row, row * 1.25, "text", "text",
                                 "text", "text", "text", "text", "text"])
    workbook.close()


for rows in (100_000, 500_000, 1_000_000):
    path = Path(tempfile.gettempdir()) / f"openpyxl-streaming-{rows}.xlsx"
    generate(path, rows)
    start = time.perf_counter()
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    cells = sum(len(row) for row in workbook["Data"].iter_rows())
    workbook.close()
    elapsed = (time.perf_counter() - start) * 1000
    print(f"BENCHMARK,python,openpyxl,streaming_{rows}_read,{elapsed:.2f},{path.stat().st_size}")
    path.unlink(missing_ok=True)
