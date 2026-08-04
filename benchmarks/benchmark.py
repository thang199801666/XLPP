#!/usr/bin/env python3
"""XL++ Benchmark Suite — Compare XLPP vs openpyxl vs xlsxwriter."""

import os, sys, time, tempfile, random, string
from pathlib import Path

# Add XLPP python binding to path
XLPP_DIR = Path(__file__).parent.parent / "bindings" / "python"
ROOT_DIR = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT_DIR))
sys.path.insert(0, str(XLPP_DIR))
import xlpp

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlsxwriter
    HAS_XLSXWRITER = True
except ImportError:
    HAS_XLSXWRITER = False

def random_text(length=20):
    return ''.join(random.choices(string.ascii_letters, k=length))

def random_float():
    return random.uniform(0, 100000)

def generate_data(rows, cols):
    data = []
    for r in range(rows):
        row = []
        for c in range(cols):
            if c == 0:
                row.append(f"Item-{r}")
            elif c == 1:
                row.append(random.randint(1, 10000))
            elif c == 2:
                row.append(round(random_float(), 2))
            else:
                row.append(random_text())
        data.append(row)
    return data

def benchmark(name, fn, *args):
    import gc; gc.collect()
    start = time.perf_counter()
    result = fn(*args)
    elapsed = time.perf_counter() - start
    return result, elapsed

def write_xlpp(path, data):
    wb = xlpp.Workbook()
    ws = wb.add_worksheet("Data")
    t0 = time.perf_counter()
    for row in data:
        ws.append(row)
    append_time = time.perf_counter() - t0
    ws["A1"].font().bold = True
    ws.freeze_panes("A2")
    t1 = time.perf_counter()
    wb.save(str(path))
    save_time = time.perf_counter() - t1
    return (append_time, save_time)

def write_openpyxl(path, data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    for row in data:
        ws.append(row)
    ws["A1"].font = openpyxl.styles.Font(bold=True)
    ws.freeze_panes = "A2"
    wb.save(str(path))

def write_xlsxwriter(path, data):
    wb = xlsxwriter.Workbook(str(path))
    ws = wb.add_worksheet("Data")
    bold = wb.add_format({"bold": True})
    for r, row in enumerate(data):
        for c, val in enumerate(row):
            if r == 0:
                ws.write(r, c, val, bold)
            else:
                ws.write(r, c, val)
    ws.freeze_panes(1, 0)
    wb.close()

def read_xlpp(path):
    wb = xlpp.Workbook()
    wb.load(str(path))
    ws = wb["Data"]
    return ws.max_row  # verify data loaded

def read_openpyxl(path):
    wb = openpyxl.load_workbook(str(path), read_only=True)
    ws = wb.active
    count = 0
    for _ in ws.iter_rows():
        count += 1
    wb.close()
    return count

def run_benchmarks():
    configs = [
        (500, 5, "500 × 5"),
        (1000, 10, "1K × 10"),
        (5000, 10, "5K × 10"),
        (10000, 15, "10K × 15"),
    ]

    results = []
    tmppath = Path(tempfile.gettempdir())

    for rows, cols, label in configs:
        print(f"\n{'='*60}")
        print(f"  Benchmark: {label}")
        print(f"{'='*60}")

        data = generate_data(rows, cols)
        row = {}

        # XLPP write
        fp = tmppath / "xlpp_bench.xlsx"
        _, elapsed = benchmark("XLPP write", write_xlpp, fp, data)
        size = fp.stat().st_size
        print(f"  XLPP   write: {elapsed*1000:>8.1f} ms  size: {size:>10,} bytes")
        row["xlpp_write_ms"] = round(elapsed * 1000, 1)
        row["xlpp_size"] = size

        # XLPP read
        _, elapsed = benchmark("XLPP read", read_xlpp, fp)
        print(f"  XLPP   read:  {elapsed*1000:>8.1f} ms")
        row["xlpp_read_ms"] = round(elapsed * 1000, 1)

        # openpyxl write
        if HAS_OPENPYXL:
            fp2 = tmppath / "openpyxl_bench.xlsx"
            _, elapsed = benchmark("openpyxl write", write_openpyxl, fp2, data)
            size2 = fp2.stat().st_size
            print(f"  openpyxl write: {elapsed*1000:>8.1f} ms  size: {size2:>10,} bytes")
            row["opy_write_ms"] = round(elapsed * 1000, 1)
            row["opy_size"] = size2

            _, elapsed = benchmark("openpyxl read", read_openpyxl, fp2)
            print(f"  openpyxl read:  {elapsed*1000:>8.1f} ms")
            row["opy_read_ms"] = round(elapsed * 1000, 1)

            os.remove(fp2)

        # xlsxwriter write
        if HAS_XLSXWRITER:
            fp3 = tmppath / "xlsxwr_bench.xlsx"
            _, elapsed = benchmark("xlsxwriter write", write_xlsxwriter, fp3, data)
            size3 = fp3.stat().st_size
            print(f"  xlsxwriter write: {elapsed*1000:>8.1f} ms  size: {size3:>10,} bytes")
            row["xwr_write_ms"] = round(elapsed * 1000, 1)
            row["xwr_size"] = size3
            os.remove(fp3)

        os.remove(fp)
        for library in ("XLPP", "openpyxl", "XlsxWriter"):
            prefix = {"XLPP": "xlpp", "openpyxl": "opy", "XlsxWriter": "xwr"}[library]
            if f"{prefix}_write_ms" in row:
                print(f"BENCHMARK,python,{library},write,{row[prefix + '_write_ms']},{row[prefix + '_size']}")
            if f"{prefix}_read_ms" in row:
                print(f"BENCHMARK,python,{library},read,{row[prefix + '_read_ms']},0")
        results.append((label, row))

    # Summary table
    print(f"\n{'='*80}")
    print("  SUMMARY")
    print(f"{'='*80}")

    header = f"{'Test':>12} | {'XLPP write':>12} | {'XLPP read':>10} | {'XLPP size':>10}"
    if HAS_OPENPYXL:
        header += f" | {'opy write':>10} | {'opy read':>10} | {'opy size':>10}"
    if HAS_XLSXWRITER:
        header += f" | {'xwr write':>10} | {'xwr size':>10}"
    print(header)
    print("-" * 100)

    for label, r in results:
        line = f"{label:>12} | {r['xlpp_write_ms']:>8.1f} ms | {r['xlpp_read_ms']:>6.1f} ms | {r['xlpp_size']:>8,} B"
        if HAS_OPENPYXL:
            line += f" | {r['opy_write_ms']:>6.1f} ms | {r['opy_read_ms']:>6.1f} ms | {r['opy_size']:>8,} B"
        if HAS_XLSXWRITER:
            line += f" | {r['xwr_write_ms']:>6.1f} ms | {r['xwr_size']:>8,} B"
        print(line)

    # Speedup
    if HAS_OPENPYXL:
        print(f"\n  XLPP vs openpyxl write speedup: {results[-1][1]['opy_write_ms']/results[-1][1]['xlpp_write_ms']:.1f}x")
        print(f"  XLPP vs openpyxl read speedup:  {results[-1][1]['opy_read_ms']/results[-1][1]['xlpp_read_ms']:.1f}x")

if __name__ == "__main__":
    run_benchmarks()
