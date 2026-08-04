"""Differential compatibility tests against openpyxl.

Generates real .xlsx files with openpyxl and verifies xlpp reads them
correctly, and vice versa. This is the seed of the roadmap's real-world
compatibility corpus (Phase 3).
"""

from datetime import date, datetime

import openpyxl
import pytest

import xlpp


# ---------------------------------------------------------------------------
# openpyxl -> xlpp
# ---------------------------------------------------------------------------

def test_basic_values(tmp_path):
    path = tmp_path / "values.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "hello"
    ws["A2"] = 42
    ws["A3"] = 3.5
    ws["A4"] = True
    ws["A5"] = None
    ws["A6"] = datetime(2024, 1, 15, 13, 30, 45)
    wb.save(path)

    x = xlpp.Workbook()
    x.load(str(path))
    assert x.sheet_names == ["Data"]
    s = x["Data"]
    assert s["A1"].value == "hello"
    assert s["A2"].value == 42.0
    assert s["A3"].value == 3.5
    assert s["A4"].value is True
    assert s["A5"].value is None
    assert s["A6"].is_date()
    assert s["A6"].value == datetime(2024, 1, 15, 13, 30, 45)


def test_styles(tmp_path):
    path = tmp_path / "styles.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Styled"
    ws["A1"].font = openpyxl.styles.Font(bold=True, italic=True, size=14, name="Arial",
                                          color="FF0000FF")
    ws["A2"] = 123
    ws["A2"].number_format = "#,##0.00"
    ws["A3"] = "Fill"
    ws["A3"].fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="FFFFFF00")
    ws["A4"] = "Border"
    ws["A4"].border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style="thin"), top=openpyxl.styles.Side(style="medium"))
    ws["A5"] = "Align"
    ws["A5"].alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center",
                                                    wrap_text=True)
    wb.save(path)

    x = xlpp.Workbook()
    x.load(str(path))
    s = x["Sheet"]
    assert s["A1"].font().bold is True
    assert s["A1"].font().italic is True
    assert s["A1"].font().size == 14.0
    assert s["A1"].font().name == "Arial"
    assert s["A1"].font().color().argb in ("FF0000FF", "000000FF")
    assert s["A2"].number_format == "#,##0.00"
    assert s["A3"].fill().pattern_type == "solid"
    assert s["A3"].fill().foreground().argb in ("FFFFFF00", "00FFFF00")
    assert s["A4"].border().left().style == "thin"
    assert s["A4"].border().top().style == "medium"
    assert s["A5"].alignment().horizontal == "center"
    assert s["A5"].alignment().vertical == "center"
    assert s["A5"].alignment().wrap_text is True


def test_dates_various_formats(tmp_path):
    path = tmp_path / "dates.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = datetime(2024, 1, 15)
    ws["A1"].number_format = "yyyy-mm-dd"
    ws["A2"] = datetime(2024, 1, 15, 14, 30, 0)
    ws["A2"].number_format = "yyyy-mm-dd h:mm:ss"
    ws["A3"] = datetime(1999, 12, 31, 23, 59, 59)
    ws["A3"].number_format = "m/d/yy"
    wb.save(path)

    x = xlpp.Workbook()
    x.load(str(path))
    s = x["Sheet"]
    assert s["A1"].is_date()
    assert s["A1"].value == datetime(2024, 1, 15)
    assert s["A2"].value == datetime(2024, 1, 15, 14, 30, 0)
    assert s["A3"].value == datetime(1999, 12, 31, 23, 59, 59)


def test_merge_and_freeze(tmp_path):
    path = tmp_path / "layout.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.merge_cells("A1:C1")
    ws["A1"] = "Title"
    ws.freeze_panes = "A2"
    wb.save(path)

    x = xlpp.Workbook()
    x.load(str(path))
    s = x["Sheet"]
    assert s.merged_ranges == ["A1:C1"]
    assert s.frozen_pane == "A2"
    assert s["A1"].value == "Title"


def test_multiple_sheets(tmp_path):
    path = tmp_path / "multi.xlsx"
    wb = openpyxl.Workbook()
    wb.create_sheet("First")
    wb.create_sheet("Second")
    wb.save(path)

    x = xlpp.Workbook()
    x.load(str(path))
    # openpyxl always keeps the default "Sheet" plus the created ones.
    assert "First" in x.sheet_names
    assert "Second" in x.sheet_names


def test_formula(tmp_path):
    path = tmp_path / "formula.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 2
    ws["A2"] = 3
    ws["A3"] = "=SUM(A1:A2)"
    wb.save(path)

    x = xlpp.Workbook()
    x.load(str(path))
    s = x["Sheet"]
    assert s["A3"].has_formula()
    assert s["A3"].formula == "SUM(A1:A2)"
    assert s["A1"].value == 2.0


def test_hyperlink(tmp_path):
    path = tmp_path / "links.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "OpenAI"
    ws["A1"].hyperlink = "https://example.com"
    wb.save(path)

    x = xlpp.Workbook()
    x.load(str(path))
    s = x["Sheet"]
    assert s["A1"].has_hyperlink()
    assert s["A1"].hyperlink().target == "https://example.com"


def test_comment(tmp_path):
    path = tmp_path / "comments.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "value"
    ws["A1"].comment = openpyxl.comments.Comment("Review this", "Alice")
    wb.save(path)

    x = xlpp.Workbook()
    x.load(str(path))
    s = x["Sheet"]
    assert s["A1"].has_comment()
    assert s["A1"].comment().text == "Review this"
    assert s["A1"].comment().author == "Alice"


def test_rich_string_with_shared_strings(tmp_path):
    path = tmp_path / "shared.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    # Repeated strings force the shared-strings table.
    for i in range(1, 21):
        ws.cell(row=i, column=1).value = "repeated"
    ws["B1"] = "unique"
    wb.save(path)

    x = xlpp.Workbook()
    x.load(str(path))
    s = x["Sheet"]
    assert s["A20"].value == "repeated"
    assert s["B1"].value == "unique"


def test_column_dimensions(tmp_path):
    path = tmp_path / "colwidth.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.column_dimensions["A"].width = 30.0
    ws["A1"] = "wide"
    wb.save(path)

    x = xlpp.Workbook()
    x.load(str(path))
    s = x["Sheet"]
    assert s["A1"].value == "wide"


# ---------------------------------------------------------------------------
# xlpp -> openpyxl
# ---------------------------------------------------------------------------

def test_xlpp_to_openpyxl_values(tmp_path):
    path = tmp_path / "xlpp_to_openpyxl.xlsx"
    x = xlpp.Workbook()
    s = x.add_worksheet("Data")
    s["A1"].value = "hello"
    s["A2"].value = 42
    s["A3"].value = 3.5
    s["A4"].value = True
    s["A5"].value = date(2024, 6, 1)
    x.save(str(path))

    wb = openpyxl.load_workbook(str(path))
    ws = wb["Data"]
    assert ws["A1"].value == "hello"
    assert ws["A2"].value == 42
    assert ws["A3"].value == 3.5
    assert ws["A4"].value is True
    assert ws["A5"].value == datetime(2024, 6, 1)


def test_xlpp_to_openpyxl_styles(tmp_path):
    path = tmp_path / "xlpp_styles.xlsx"
    x = xlpp.Workbook()
    s = x.add_worksheet("Styled")
    s["A1"].value = "Bold"
    f = s["A1"].font()
    f.bold = True
    f.size = 16.0
    s["A1"].set_number_format("0.00")
    s["B1"].value = 0
    s["B1"].fill().pattern_type = "solid"
    s["B1"].fill().foreground().set_argb("FFFFFF00")
    s["B1"].border().top().style = "medium"
    x.save(str(path))

    wb = openpyxl.load_workbook(str(path))
    ws = wb["Styled"]
    assert ws["A1"].font.bold is True
    assert ws["A1"].font.size == 16.0
    assert ws["A1"].number_format == "0.00"
    assert ws["B1"].fill.patternType == "solid"
    assert ws["B1"].border.top.style == "medium"


def test_xlpp_to_openpyxl_roundtrip_merge(tmp_path):
    path = tmp_path / "xlpp_merge.xlsx"
    x = xlpp.Workbook()
    s = x.add_worksheet("Layout")
    s["A1"].value = "Title"
    s.merge_cells("A1:C1")
    s.freeze_panes("A2")
    x.save(str(path))

    wb = openpyxl.load_workbook(str(path))
    ws = wb["Layout"]
    assert "A1:C1" in [str(r) for r in ws.merged_cells.ranges]
    assert ws.freeze_panes == "A2"
    assert ws["A1"].value == "Title"


def test_openpyxl_xlpp_full_roundtrip(tmp_path):
    """openpyxl writes -> xlpp loads -> xlpp saves -> openpyxl reads."""
    first = tmp_path / "stage1.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "name"
    ws["B1"] = "amount"
    ws["A2"] = "Widget"
    ws["B2"] = 19.99
    ws["A2"].font = openpyxl.styles.Font(bold=True)
    wb.save(first)

    x = xlpp.Workbook()
    x.load(str(first))
    s = x["Sheet"]
    assert s["A2"].value == "Widget"
    assert s["A2"].font().bold is True

    second = tmp_path / "stage2.xlsx"
    x.save(str(second))

    wb2 = openpyxl.load_workbook(str(second))
    ws2 = wb2["Sheet"]
    assert ws2["A2"].value == "Widget"
    assert ws2["B2"].value == 19.99
    assert ws2["A2"].font.bold is True


def test_openpyxl_empty_cells_roundtrip(tmp_path):
    path = tmp_path / "empty.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "start"
    ws["C3"] = "end"
    wb.save(path)

    x = xlpp.Workbook()
    x.load(str(path))
    s = x["Sheet"]
    assert s["A1"].value == "start"
    assert s["C3"].value == "end"


def test_special_characters_openpyxl_to_xlpp(tmp_path):
    path = tmp_path / "special.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "a < b & c > d"
    ws["A2"] = 'quotes "and" \'single\''
    ws["A3"] = "line1\nline2"
    ws["A4"] = "  leading/trailing  "
    wb.save(path)

    x = xlpp.Workbook()
    x.load(str(path))
    s = x["Sheet"]
    assert s["A1"].value == "a < b & c > d"
    assert s["A2"].value == 'quotes "and" \'single\''
    assert s["A3"].value == "line1\nline2"
    assert s["A4"].value == "  leading/trailing  "


def test_special_characters_xlpp_to_openpyxl(tmp_path):
    path = tmp_path / "special_out.xlsx"
    x = xlpp.Workbook()
    s = x.add_worksheet("Data")
    s["A1"].value = "a < b & c > d"
    s["A2"].value = 'quotes "and" \'single\''
    x.save(str(path))

    wb = openpyxl.load_workbook(str(path))
    ws = wb["Data"]
    assert ws["A1"].value == "a < b & c > d"
    assert ws["A2"].value == 'quotes "and" \'single\''


def test_autofilter_openpyxl(tmp_path):
    path = tmp_path / "filter.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Name", "Status"])
    ws.append(["Alpha", "Open"])
    ws.append(["Beta", "Closed"])
    ws.auto_filter.ref = "A1:B3"
    wb.save(path)

    x = xlpp.Workbook()
    x.load(str(path))
    s = x["Sheet"]
    assert s.auto_filter().reference == "A1:B3"


def test_large_and_extreme_values(tmp_path):
    path = tmp_path / "numbers.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 123456789012345.0
    ws["A2"] = -0.001
    ws["A3"] = 1e-10
    ws["A4"] = 1e30
    wb.save(path)

    x = xlpp.Workbook()
    x.load(str(path))
    s = x["Sheet"]
    assert s["A1"].value == 123456789012345.0
    assert s["A2"].value == -0.001
    assert s["A3"].value == 1e-10
    assert s["A4"].value == 1e30


def test_mixed_types_full_roundtrip(tmp_path):
    """A realistic mixed workbook through openpyxl -> xlpp -> openpyxl."""
    first = tmp_path / "mixed1.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Product"
    ws["B1"] = "Price"
    ws["C1"] = "In Stock"
    ws["A2"] = "Widget"
    ws["B2"] = 19.99
    ws["C2"] = True
    ws["A3"] = "Gadget"
    ws["B3"] = 4.5
    ws["C3"] = False
    ws["B2"].number_format = "0.00"
    ws["A2"].font = openpyxl.styles.Font(bold=True)
    wb.save(first)

    x = xlpp.Workbook()
    x.load(str(first))
    s = x["Sheet"]
    assert s["A2"].value == "Widget"
    assert s["B2"].value == 19.99
    assert s["C2"].value is True
    assert s["B2"].number_format == "0.00"
    assert s["A2"].font().bold is True

    second = tmp_path / "mixed2.xlsx"
    x.save(str(second))

    wb2 = openpyxl.load_workbook(str(second))
    ws2 = wb2["Sheet"]
    assert ws2["A3"].value == "Gadget"
    assert ws2["B3"].value == 4.5
    assert ws2["C3"].value is False
    assert ws2["A2"].font.bold is True
    assert ws2["B2"].number_format == "0.00"


def test_number_formats(tmp_path):
    path = tmp_path / "fmt.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 0.5
    ws["A1"].number_format = "0.00%"
    ws["A2"] = 1234.5
    ws["A2"].number_format = "$#,##0.00"
    ws["A3"] = 12345.6789
    ws["A3"].number_format = "0.00E+00"
    ws["A4"] = 3.14159
    ws["A4"].number_format = "0.0000"
    ws["A5"] = 42
    ws["A5"].number_format = "0"
    wb.save(path)

    x = xlpp.Workbook()
    x.load(str(path))
    s = x["Sheet"]
    assert s["A1"].number_format == "0.00%"
    assert s["A2"].number_format == "$#,##0.00"
    assert s["A3"].number_format == "0.00E+00"
    assert s["A4"].number_format == "0.0000"
    assert s["A5"].number_format == "0"
    assert s["A1"].value == 0.5


def test_formula_cached_value(tmp_path):
    path = tmp_path / "cached.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 2
    ws["A2"] = 3
    ws["A3"] = "=A1*A2"
    ws["A4"] = "=SUM(A1:A2)"
    wb.save(path)

    x = xlpp.Workbook()
    x.load(str(path))
    s = x["Sheet"]
    assert s["A3"].has_formula()
    assert s["A3"].formula == "A1*A2"
    assert s["A4"].has_formula()
    assert s["A4"].formula == "SUM(A1:A2)"


def test_error_cells(tmp_path):
    path = tmp_path / "errors.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "=1/0"
    ws["A2"] = "=NOTHING"
    wb.save(path)

    x = xlpp.Workbook()
    x.load(str(path))
    s = x["Sheet"]
    assert s["A1"].has_formula()
    assert s["A2"].has_formula()


def test_date1904_from_openpyxl(tmp_path):
    path = tmp_path / "d1904.xlsx"
    wb = openpyxl.Workbook()
    wb.epoch = openpyxl.utils.datetime.CALENDAR_MAC_1904
    ws = wb.active
    ws["A1"] = datetime(2024, 1, 15)
    ws["A1"].number_format = "yyyy-mm-dd"
    wb.save(path)

    x = xlpp.Workbook()
    x.load(str(path))
    assert x.date_1904 is True
    s = x["Sheet"]
    assert s["A1"].is_date()
    assert s["A1"].value == datetime(2024, 1, 15)


def test_large_sheet(tmp_path):
    path = tmp_path / "large.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    for r in range(1, 501):
        ws.cell(row=r, column=1).value = f"row{r}"
        ws.cell(row=r, column=2).value = r
    wb.save(path)

    x = xlpp.Workbook()
    x.load(str(path))
    s = x["Sheet"]
    assert s["A1"].value == "row1"
    assert s["B250"].value == 250.0
    assert s["A500"].value == "row500"
    assert s["B500"].value == 500.0
    assert s.max_row == 500


def test_freeze_and_filter_combined(tmp_path):
    path = tmp_path / "combo.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["A", "B", "C"])
    ws.append([1, 2, 3])
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = "A1:C2"
    wb.save(path)

    x = xlpp.Workbook()
    x.load(str(path))
    s = x["Sheet"]
    assert s.frozen_pane == "B2"
    assert s.auto_filter().reference == "A1:C2"
    assert s["B2"].value == 2.0


def test_builtin_number_format_id_3(tmp_path):
    """openpyxl uses built-in numFmtId 3 for '#,##0' without declaring it."""
    path = tmp_path / "id3.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = 1234
    ws["A1"].number_format = "#,##0"
    ws["A2"] = 1234.56
    ws["A2"].number_format = "#,##0.00"
    wb.save(path)

    x = xlpp.Workbook()
    x.load(str(path))
    s = x["Sheet"]
    assert s["A1"].number_format == "#,##0"
    assert s["A2"].number_format == "#,##0.00"


def test_styled_empty_cell_preserved(tmp_path):
    """A styled cell with no value must survive save/load (openpyxl writes it)."""
    path = tmp_path / "styled_empty.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "x"
    ws["B2"].fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="FFFFFF00")
    wb.save(path)

    x = xlpp.Workbook()
    x.load(str(path))
    s = x["Sheet"]
    assert s["B2"].fill().pattern_type == "solid"

    out = tmp_path / "styled_empty_out.xlsx"
    x.save(str(out))

    wb2 = openpyxl.load_workbook(str(out))
    assert wb2["Sheet"]["B2"].fill.patternType == "solid"


def test_styled_empty_cell_from_xlpp(tmp_path):
    """xlpp-created styled-empty cells must be serialized."""
    path = tmp_path / "styled_empty_xlpp.xlsx"
    x = xlpp.Workbook()
    s = x.add_worksheet("S")
    s["A1"].value = "x"
    s["B2"].fill().pattern_type = "solid"
    s["B2"].fill().foreground().set_argb("FFFFFF00")
    x.save(str(path))

    wb = openpyxl.load_workbook(str(path))
    assert wb["S"]["B2"].fill.patternType == "solid"
