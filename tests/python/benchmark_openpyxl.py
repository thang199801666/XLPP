"""Small, reproducible XL++ versus OpenPyXL benchmark."""

import argparse
import statistics
import tempfile
import time
from pathlib import Path

import xlpp


def benchmark_xlpp(path, rows, columns):
    book = xlpp.Workbook()
    sheet = book.add_worksheet("Data")
    start = time.perf_counter()
    sheet.append([f"C{i}" for i in range(columns)])
    for row in range(rows):
        sheet.append([row * columns + col for col in range(columns)])
    write_start = time.perf_counter()
    book.save(str(path))
    write_seconds = time.perf_counter() - write_start

    loaded = xlpp.Workbook()
    read_start = time.perf_counter()
    loaded.load(str(path))
    read_seconds = time.perf_counter() - read_start
    return time.perf_counter() - start, write_seconds, read_seconds


def benchmark_openpyxl(path, rows, columns):
    from openpyxl import Workbook, load_workbook

    start = time.perf_counter()
    book = Workbook(write_only=False)
    sheet = book.active
    sheet.title = "Data"
    sheet.append([f"C{i}" for i in range(columns)])
    for row in range(rows):
        sheet.append([row * columns + col for col in range(columns)])
    write_start = time.perf_counter()
    book.save(path)
    write_seconds = time.perf_counter() - write_start
    read_start = time.perf_counter()
    load_workbook(path, read_only=True).close()
    read_seconds = time.perf_counter() - read_start
    return time.perf_counter() - start, write_seconds, read_seconds


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--rows", type=int, default=10000)
    parser.add_argument("--columns", type=int, default=10)
    args = parser.parse_args()
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        xlpp_result = benchmark_xlpp(root / "xlpp.xlsx", args.rows, args.columns)
        openpyxl_result = benchmark_openpyxl(str(root / "openpyxl.xlsx"), args.rows, args.columns)
    print(f"rows={args.rows} columns={args.columns}")
    print(f"XL++     total={xlpp_result[0]:.3f}s write={xlpp_result[1]:.3f}s read={xlpp_result[2]:.3f}s")
    print(f"openpyxl total={openpyxl_result[0]:.3f}s write={openpyxl_result[1]:.3f}s read={openpyxl_result[2]:.3f}s")
    print(f"write speedup={openpyxl_result[1] / xlpp_result[1]:.2f}x")
    print(f"read speedup={openpyxl_result[2] / xlpp_result[2]:.2f}x")


if __name__ == "__main__":
    main()
