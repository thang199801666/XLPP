#!/usr/bin/env python3
"""Optional P1S interoperability check against an installed openpyxl.

This is intentionally not part of the default CTest graph because openpyxl is
an external test dependency. Pass artifacts produced by
XLPP_P1S_ThreePillarTests with XLPP_KEEP_P1S_ARTIFACTS=1.
"""
from __future__ import annotations

import argparse
from pathlib import Path
import sys


def require(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--template", type=Path, required=True)
    parser.add_argument("--macro-template", type=Path, required=True)
    parser.add_argument("--charts", type=Path, required=True)
    parser.add_argument("--pivot", type=Path, required=True)
    args = parser.parse_args()

    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit("openpyxl is required for this optional host check") from exc

    template = openpyxl.load_workbook(args.template)
    require(template.template is True, "XLTX was not recognized as a template")

    macro_template = openpyxl.load_workbook(args.macro_template, keep_vba=True)
    require(macro_template.template is True, "XLTM was not recognized as a template")

    chart_book = openpyxl.load_workbook(args.charts)
    ws = chart_book["Data"]
    require(len(ws._charts) == 3, "expected three generated charts")
    require(type(ws._charts[0]).__name__ == "ScatterChart", "scatter chart type mismatch")
    require(type(ws._charts[1]).__name__ == "BubbleChart", "bubble chart type mismatch")
    combo = ws._charts[2]
    plots = list(getattr(combo, "_charts", [combo]))
    require([type(c).__name__ for c in plots] == ["BarChart", "LineChart"],
            "combined chart was not recognized as Bar + Line")
    require(list(plots[0].axId) == [10, 100], "primary plot axis IDs mismatch")
    require(list(plots[1].axId) == [10, 200], "secondary plot axis IDs mismatch")

    pivot_book = openpyxl.load_workbook(args.pivot)
    pivot_ws = pivot_book["Data"]
    require(len(pivot_ws._pivots) == 1, "generated PivotTable was not discovered")

    print(f"P1S openpyxl host check: PASS (openpyxl {openpyxl.__version__})")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except AssertionError as exc:
        print(f"P1S openpyxl host check: FAIL: {exc}", file=sys.stderr)
        raise SystemExit(1)
