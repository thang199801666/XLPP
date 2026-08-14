#!/usr/bin/env python3
"""P1V external-host gate for advanced Chartsheet semantics.

Usage:
  python p1v_openpyxl_chartsheet_depth_check.py <advanced.xltx> <custom-mutated.xltx> <worksheet.xlsx>
"""
from pathlib import Path
import sys
from openpyxl import load_workbook


def require(cond, label):
    if not cond:
        raise SystemExit(f"P1V openpyxl host check failed: {label}")


def check_advanced(path: Path):
    wb = load_workbook(path)
    require(wb.template is True, "XLTX template identity")
    require(wb.active.title == "Dashboard" and len(wb.chartsheets) == 1, "active Chartsheet")
    cs = wb.chartsheets[0]
    ps = cs.pageSetup
    require(ps.paperHeight == "210mm" and ps.paperWidth == "297mm", "paper dimensions")
    require(ps.pageOrder == "overThenDown" and ps.usePrinterDefaults is False, "page order/printer defaults")
    require(ps.cellComments == "atEnd" and ps.errors == "dash", "comment/error print behavior")
    require(ps.horizontalDpi == 600 and ps.verticalDpi == 600 and ps.copies == 3, "DPI/copies")
    protection = cs.sheetProtection
    require(protection.algorithmName == "SHA-512" and protection.hashValue == "AQIDBA==", "modern protection hash metadata")
    require(protection.saltValue == "BQYHCA==" and protection.spinCount == 100000, "modern protection salt/spins")
    require(cs.customSheetViews is not None and len(cs.customSheetViews.customSheetView) == 1, "custom chart-sheet view")
    view = cs.customSheetViews.customSheetView[0]
    require(view.guid == "{11111111-2222-3333-4444-555555555555}", "custom view GUID")
    require(view.scale == 90 and view.state == "hidden" and view.zoomToFit is False, "custom view state")
    require(view.pageSetup.horizontalDpi == 300 and view.pageSetup.copies == 2, "custom view nested setup")
    require(view.headerFooter.oddHeader.center.text == "Custom View", "custom view nested header")
    wb.close()


def check_custom_mutated(path: Path):
    wb = load_workbook(path)
    cs = wb.chartsheets[0]
    require(cs.customSheetViews.customSheetView[0].scale == 95, "mutated custom view scale")
    wb.close()


def check_worksheet(path: Path):
    wb = load_workbook(path)
    ps = wb["Sheet1"].page_setup
    require(ps.paperHeight == "11in" and ps.paperWidth == "8.5in", "worksheet paper dimensions")
    require(ps.pageOrder == "downThenOver" and ps.usePrinterDefaults is True, "worksheet extended page setup")
    require(ps.horizontalDpi == 1200 and ps.verticalDpi == 1200 and ps.copies == 4, "worksheet DPI/copies")
    wb.close()


if __name__ == "__main__":
    if len(sys.argv) != 4:
        raise SystemExit(__doc__)
    check_advanced(Path(sys.argv[1]))
    check_custom_mutated(Path(sys.argv[2]))
    check_worksheet(Path(sys.argv[3]))
    print("P1V openpyxl Chartsheet-depth host check: PASS")
