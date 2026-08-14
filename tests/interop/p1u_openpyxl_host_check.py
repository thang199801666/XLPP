#!/usr/bin/env python3
"""P1U external-host gate for XLTX/XLTM and Chartsheet semantics.

Usage:
  python p1u_openpyxl_host_check.py <mixed.xltx> <patched_import.xltx> <macro.xltm>
"""
from pathlib import Path
from zipfile import ZipFile
import sys
from openpyxl import load_workbook


def require(cond, label):
    if not cond:
        raise SystemExit(f"P1U openpyxl host check failed: {label}")


def check_mixed(path: Path):
    wb = load_workbook(path)
    require(wb.template is True, "XLTX template identity")
    require(wb.sheetnames == ["Data", "Dashboard", "Hidden Data", "Secret Chart"], "mixed tab order")
    require(wb.active.title == "Dashboard", "active Chartsheet")
    require([cs.title for cs in wb.chartsheets] == ["Dashboard", "Secret Chart"], "Chartsheet discovery")
    require(wb["Hidden Data"].sheet_state == "hidden", "hidden worksheet state")
    cs = wb.chartsheets[0]
    require(cs.sheetViews.sheetView[0].zoomScale == 125 and cs.sheetViews.sheetView[0].zoomToFit is False, "Chartsheet view")
    require(cs.sheetPr.codeName == "ChartDashboard" and cs.sheetPr.tabColor.rgb == "FF336699", "Chartsheet properties")
    require(cs.sheetProtection is not None and cs.sheetProtection.content and cs.sheetProtection.objects, "Chartsheet protection")
    require(cs.pageSetup.orientation == "landscape" and cs.pageSetup.scale == 85, "Chartsheet page setup")
    require(cs.headerFooter.oddHeader.center.text == "Advanced Header", "Chartsheet header")
    # openpyxl 3.1.x currently does not copy workbook manifest state back to
    # Chartsheet.sheet_state on read. Check the OOXML manifest independently.
    with ZipFile(path) as z:
        workbook_xml = z.read("xl/workbook.xml").decode("utf-8")
    require('name="Secret Chart"' in workbook_xml and 'state="veryHidden"' in workbook_xml, "veryHidden Chartsheet manifest state")
    wb.close()


def check_patched(path: Path):
    wb = load_workbook(path)
    require(wb.template is True and wb.active.title == "Dashboard", "patched imported template identity")
    cs = wb.chartsheets[0]
    require(cs.sheetViews.sheetView[0].zoomScale == 140, "patched Chartsheet zoom")
    require(cs.pageSetup.scale == 90, "patched Chartsheet page setup")
    require(cs.headerFooter.oddHeader.center.text == "Patched Header", "patched Chartsheet header")
    require(cs.headerFooter.firstHeader.left.text == "First Header", "first-page header preservation")
    require(len(cs._charts) == 1, "imported chart ownership remains readable")
    wb.close()


def check_macro(path: Path):
    require(path.suffix.lower() == ".xltm", "XLTM file extension")
    with ZipFile(path) as z:
        require("xl/vbaProject.bin" in z.namelist(), "XLTM contains VBA project part")
        content_types = z.read("[Content_Types].xml").decode("utf-8")
    require("application/vnd.ms-excel.template.macroEnabled.main+xml" in content_types, "XLTM macro-template content type")
    wb = load_workbook(path, keep_vba=True)
    require(wb.template is True, "XLTM template identity")
    require(wb.vba_archive is not None, "XLTM VBA archive")
    require(wb.active.title == "Dashboard" and len(wb.chartsheets) == 1, "XLTM active Chartsheet")
    wb.close()


if __name__ == "__main__":
    if len(sys.argv) != 4:
        raise SystemExit(__doc__)
    check_mixed(Path(sys.argv[1]))
    check_patched(Path(sys.argv[2]))
    check_macro(Path(sys.argv[3]))
    print("P1U openpyxl host check: PASS")
